import os
import sys
import time
import json
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# 🔧 환경 초기화
now = datetime.now()
today_str = now.strftime('%Y-%m-%d')
timestamp = now.strftime('%Y%m%d_%H%M%S')
log_dir = f"./logs/{today_str}"
result_dir = f"./result/{today_str}"
os.makedirs(log_dir, exist_ok=True)
os.makedirs(result_dir, exist_ok=True)
log_path = f"{log_dir}/log_{timestamp}.txt"
config_path = "config.json"

# 🔧 로그 이중 출력
class DualLogger:
    def __init__(self, path):
        self.terminal = sys.stdout
        self.log = open(path, "w", encoding="utf-8")
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    def flush(self):
        self.terminal.flush()
        self.log.flush()
sys.stdout = DualLogger(log_path)
# 🔧 브라우저 세팅
chrome_driver_path = "./chromedriver.exe"
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(chrome_driver_path), options=options)
wait = WebDriverWait(driver, 20)

# 🔧 키워드 로딩
with open("keywords.txt", encoding="utf-8") as f:
    keywords = [line.strip() for line in f.readlines() if line.strip()]

# 🔧 환경 설정 로드
def load_last_run_time():
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            try:
                config = json.load(f)
                return datetime.strptime(config.get("last_run_time", ""), "%Y%m%d%H%M%S")
            except Exception as e:
                print(f"⚠️ 설정파일 파싱 오류: {e}")
    return now.replace(day=1)

def save_last_run_time():
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump({"last_run_time": now.strftime("%Y%m%d%H%M%S")}, f, indent=2)

from_bid_dt = load_last_run_time().strftime("%Y%m%d")
to_bid_dt = now.strftime("%Y%m%d")
# 🔍 팝업 제거
def force_remove_popups(driver):
    popup_selectors = [
        "div.popup", "div.popLayer", "div.pop_wrap", "div.w2window", "div.banner-wrap"
    ]
    for selector in popup_selectors:
        try:
            for div in driver.find_elements(By.CSS_SELECTOR, selector):
                if div.is_displayed():
                    driver.execute_script("arguments[0].remove()", div)
        except:
            continue

# 📄 시트 스타일 정리
def format_worksheet(ws):
    font = Font(name='맑은 고딕', size=10)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

try:
    driver.get("https://www.g2b.go.kr")
    time.sleep(8)
    force_remove_popups(driver)

    # 입찰공고 목록 클릭
    menus = driver.find_elements(By.XPATH, "//span[text()='입찰공고목록']")
    if menus:
        driver.execute_script("arguments[0].click();", menus[0])
        print("✅ 입찰공고목록 진입 완료")
    time.sleep(5)

    # 페이지당 출력 수 100으로 변경
    try:
        dropdown = wait.until(EC.presence_of_element_located((By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_sbxRecordCountPerPage1")))
        driver.execute_script("arguments[0].value = '100'; arguments[0].dispatchEvent(new Event('change'));", dropdown)
        time.sleep(3)
    except:
        print("⚠️ 리스트 수 조정 실패")

    input_box = wait.until(EC.presence_of_element_located((By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_bidPbancNm")))
    search_btn = wait.until(EC.element_to_be_clickable((By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_btnS0004")))
    start_date = driver.find_element(By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_rbxPblntfDetailDt_input_0")
    end_date = driver.find_element(By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_rbxPblntfDetailDt_input_1")

    start_date.clear()
    start_date.send_keys(from_bid_dt)
    end_date.clear()
    end_date.send_keys(to_bid_dt)

    results_dict = {}

    for keyword in keywords:
        print(f"\n🔍 검색어: {keyword}")
        input_box.clear()
        input_box.send_keys(keyword)
        search_btn.click()
        time.sleep(4)

        rows = driver.find_elements(By.CSS_SELECTOR, "tbody[id$='_body_gridView1_body_tbody'] > tr")
        print(f"🔎 검색결과 행 수: {len(rows)}")
        results = []

        for row in rows:
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 10: continue
                title = cells[6].text.strip()
                bid_no = cells[5].text.strip()
                category = cells[1].text.strip()
                org = cells[7].text.strip()
                client = cells[8].text.strip()
                post_info = cells[9].text.strip()
                post_dt, close_dt = post_info.split("\n") if "\n" in post_info else (post_info, "")
                summary_btn = row.find_element(By.CSS_SELECTOR, "td[col_id='summary'] button")
                driver.execute_script("arguments[0].scrollIntoView(true);", summary_btn)
                webdriver.ActionChains(driver).move_to_element(summary_btn).perform()
                time.sleep(1)
                tooltip = driver.find_element(By.CLASS_NAME, "w2grid_tooltip").text

                def extract_tooltip(label):
                    for line in tooltip.split("\n"):
                        if label in line:
                            return line.split(":", 1)[1].strip().replace("원", "").replace(",", "")
                    return ""

                result = {
                    "키워드": keyword,
                    "공고명": title,
                    "공고번호": bid_no,
                    "용역구분": category,
                    "공고기관": org,
                    "수요기관": client,
                    "게시일시": post_dt.strip(),
                    "입찰마감일시": close_dt.replace("(", "").replace(")", "").strip(),
                    "배정예산": extract_tooltip("배정예산"),
                    "추정가격": extract_tooltip("추정가격"),
                    "낙찰방법": extract_tooltip("낙찰방법"),
                    "재입찰": extract_tooltip("재입찰"),
                    "국내/국제 입찰사유": extract_tooltip("입찰사유"),
                    "예가방법": extract_tooltip("예가방법"),
                    "추첨번호공개여부": extract_tooltip("추첨번호공개여부"),
                }
                results.append(result)
            except Exception as e:
                print(f"⚠️ 파싱 오류: {e}")
                continue

        if results:
            results_dict[keyword] = results
        else:
            print(f"❗ 검색결과 없음: {keyword}")

    # 엑셀 파일로 저장 (시트 분할)
    xlsx_path = f"{result_dir}/입찰공고검색_{timestamp}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        for keyword, rows in results_dict.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, index=False, sheet_name=keyword[:30])  # 시트명 31자 제한

    # 서식 적용
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        format_worksheet(ws)
    wb.save(xlsx_path)
    print(f"📁 저장 완료: {xlsx_path}")

    save_last_run_time()

except Exception as e:
    print(f"❌ 전체 오류 발생: {e}")
finally:
    driver.quit()
    print("✅ 브라우저 종료 완료")
