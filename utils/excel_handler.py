# utils/excel_handler.py
# version: 1.1.0
# changelog:
# - v1.0.0: 엑셀 서식 적용 저장 기능
# - v1.1.0: 키워드별 시트 저장 방식 적용

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def save_to_excel(results_by_keyword, save_dir, filename):
    os.makedirs(save_dir, exist_ok=True)
    filepath = os.path.join(save_dir, filename)

    if not results_by_keyword:
        print("⚠️ 저장할 데이터가 없습니다.")
        return

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for keyword, records in results_by_keyword.items():
            if not records:
                continue
            df = pd.DataFrame(records)
            sheet_name = keyword[:31]
            df.to_excel(writer, index=False, sheet_name=sheet_name)

    # 서식 적용
    wb = load_workbook(filepath)
    font = Font(name="맑은 고딕", size=10)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filepath)
    print(f"📁 엑셀 저장 완료: {filepath}")
