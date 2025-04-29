import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def save_to_excel(all_results, save_dir, save_filename):
    os.makedirs(save_dir, exist_ok=True)
    filepath = os.path.join(save_dir, save_filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for keyword, data in all_results.items():
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=keyword[:31], index=False)

        writer.save()

    # 스타일 적용
    wb = load_workbook(filepath)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.font = Font(size=10)  # 폰트 10 포인트로 축소
        ws.column_dimensions["A"].auto_size = True
    wb.save(filepath)
    print(f"📁 결과 저장 완료: {filepath}")
