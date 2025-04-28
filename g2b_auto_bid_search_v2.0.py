import os
import sys
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ────────────────────────────────
# 🔧 로그 경로 설정
now = datetime.now()
log_dir = f"./logs/{now.strftime('%Y-%m-%d')}"
os.makedirs(log_dir, exist_ok=True)
log_path = f"{log_dir}/log_{now.strftime('%Y%m%d_%H%M%S')}.txt"

class DualLogger:
    def __init__(self, path):
        self.terminal = sys.stdout
        self.log = open(path, "w", encoding="utf-8")
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    def flush(self):
        self.terminal.flush()
        self.log.flush()

sys.stdout = DualLogger(log_path)

# ────────────────────────────────
# 🔧 Selenium 기본 셋업
chrome_driver_path = "./chromedriver.exe"
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(chrome_driver_path), options=options)
wait = WebDriverWait(driver, 20)
actions = ActionChains(driver)

def force_remove_popups(driver):
    selectors = ["div.popup", "div.popLayer", "div.pop_wrap", "div.w2window", "div.banner-wrap"]
    print("🔍 팝업 제거 시도")
    for selector in selectors:
        try:
            for el in driver.find_elements(By.CSS_SELECTOR, selector):
                if el.is_displayed():
                    driver.execute_script("arguments[0].remove()", el)
                    print(f"✅ 팝업 제거됨: {selector}")
        except:
            continue

# ────────────────────────────────
# 🔍 키워드 불러오기
with open("keywords.txt", encoding="utf-8") as f:
    keywords = [line.strip() for line in f.readlines() if line.strip()]

try:
    driver.get("https://www.g2b.go.kr")
    print("🌐 사이트 접속 완료")
    time.sleep(8)
    force_remove_popups(driver)

    # 메뉴 클릭
    menu = driver.find_elements(By.XPATH, "//span[text()='입찰공고목록']")
    if menu:
        driver.execute_script("arguments[0].click();", menu[0])
        print("✅ 입찰공고목록 클릭 완료")

    time.sleep(5)

    # 검색 영역 및 셀렉트박스 처리
    input_box = wait.until(EC.presence_of_element_located((By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_bidPbancNm")))
    search_btn = wait.until(EC.element_to_be_clickable((By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_btnS0004")))

    # 행 개수 설정
    select_el = driver.find_element(By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_sbxRecordCountPerPage1")
    driver.execute_script("arguments[0].value = '100'; arguments[0].dispatchEvent(new Event('change'));", select_el)
    print("📄 행 개수 100건 설정")

    print("✅ 검색 UI 준비 완료")
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for keyword in keywords:
        print(f"\n🔍 검색어: {keyword}")
        input_box.clear()
        input_box.send_keys(keyword)
        search_btn.click()
        time.sleep(3)

        rows = driver.find_elements(By.CSS_SELECTOR, "tbody[id$='_body_gridView1_body_tbody'] > tr")
        print(f"🔎 검색결과 행 수: {len(rows)}")

        results = []
        for idx, row in enumerate(rows):
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 13:
                    continue

                # 툴팁 금액 정보 확보
                summary_cell = row.find_element(By.CSS_SELECTOR, "td[col_id='summary']")
                actions.move_to_element(summary_cell).perform()
                time.sleep(0.5)

                tooltip = driver.find_element(By.ID, "mf_wfm_container_tacBidPbancLst_contents_tab2_body_gridView1_tooltip")
                tooltip_text = tooltip.text if tooltip.is_displayed() else ""

                def extract_tooltip_value(label):
                    for line in tooltip_text.split("\n"):
                        if label in line:
                            return line.split(" : ")[-1].replace("원", "").replace(",", "").strip()
                    return ""

                result = {
                    "키워드": keyword,
                    "공고명": cells[6].text.strip(),
                    "공고번호": cells[5].text.strip(),
                    "용역구분": cells[1].text.strip(),
                    "공고기관": cells[7].text.strip(),
                    "수요기관": cells[8].text.strip(),
                    "게시일시": cells[9].text.split("\n")[0].strip(),
                    "입찰마감일시": cells[9].text.split("\n")[1].replace("(", "").replace(")", "").strip() if "\n" in cells[9].text else "",
                    "배정예산": extract_tooltip_value("배정예산"),
                    "추정가격": extract_tooltip_value("추정가격"),
                    "낙찰방법": extract_tooltip_value("낙찰방법"),
                    "재입찰": extract_tooltip_value("재입찰"),
                    "국내/국제 입찰사유": extract_tooltip_value("입찰사유"),
                    "예가방법": extract_tooltip_value("예가방법"),
                    "추첨번호공개여부": extract_tooltip_value("추첨번호공개여부")
                }

                results.append(result)
            except Exception as e:
                print(f"⚠️ 행 파싱 오류: {e}")
                continue

        # 시트 작성
        if results:
            df = pd.DataFrame(results)
            sheet_name = keyword[:31].replace("/", "_").replace(":", "_")  # 엑셀 시트 제한 고려
            ws = wb.create_sheet(title=sheet_name)

            for r_idx, row in enumerate(df.itertuples(index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx+1, column=c_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.font = Font(size=10)

            # 헤더
            for col_idx, col in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col).font = Font(bold=True)
                ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', wrap_text=True)

            # 열 너비 조정
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max_len * 1.2, 50)

    # 엑셀 저장
    result_dir = f"./result/{now.strftime('%Y-%m-%d')}"
    os.makedirs(result_dir, exist_ok=True)
    result_path = f"{result_dir}/입찰공고_통합_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(result_path)
    print(f"\n📁 모든 시트 저장 완료: {result_path}")

finally:
    driver.quit()
    print("✅ 브라우저 종료 완료")
